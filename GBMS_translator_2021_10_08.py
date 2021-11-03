#Python3.9
#Program reads xml export from SUPPLIER online configurator. Then compares SUPLIER part numbers
#to COMPANY parts and creates a list of parts to be put into CRM with quantities
import openpyxl, bs4, os
import tkinter as tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
from openpyxl.styles import Font

def open_file():
    """Open a file for editing."""
    global filepath                     #has to be global so execute functions reads and accepts it
    filepath = askopenfilename(filetypes=[("XML Files", "*.xml")])
    if not filepath:
        return
    with open(filepath, "r") as input_file:
        text = input_file.read()
    window.title(f"part translator - {filepath}")

def execute():
    """Open configuration xml and look for parts_list from A cell"""
    try:                                # in case no file is selected prompts message in red
        supplier_base = open(filepath)
        soup = bs4.BeautifulSoup(supplier_base)
        supp_parts = soup.select("id")  # selected tags from xml as tag objects
        supp_parts_clean = list()  # clean list of parts as strings to be put to output excel
        qty = soup.select("count")  # selected tags from xml as tag objects
        qty_clean = list()  # clean qty list as int to be put to output excel
        for part in supp_parts:
            supp_parts_clean.append(part.text)
        for q in qty:
            qty_clean.append(int(q.text))
        COMPANY_parts = list()  # COMPANY parts corresponding to list

        # lookup parts_list to find COMPANY P/N related to SUPPLIER part number
        for COMPANY_part in supp_parts_clean:  # iterate over each part in supplier parts from xml
            COMPANY_parts.append(parts_list.get(COMPANY_part))

        # create new excel and export [supp part, company part, qty] in each column
        export_excel = openpyxl.Workbook()  # new excel file
        exp_sheet = export_excel.active
        exp_sheet.cell(row=1, column=1).value = "SUPPLIER P/N"
        exp_sheet.cell(row=1, column=2).value = "COMPANY P/N"
        exp_sheet.cell(row=1, column=3).value = "Quantity"
        exp_sheet.cell(row=1, column=4).value = "Purchasing price total €"
        exp_sheet.cell(row=1, column=6).value = "Additional options"
        exp_sheet.cell(row=2, column=6).value = "Engineering"
        exp_sheet.cell(row=3, column=6).value = "Drawing"
        exp_sheet.cell(row=4, column=6).value = "Support fee"
        exp_sheet.cell(row=5, column=6).value = "DELIVERY"
        exp_sheet.cell(row=6, column=6).value = "SHIPPING HANDLING"
        exp_sheet.cell(row=7, column=6).value = "Packing"
        exp_sheet.cell(row=1, column=7).value = "Option price €"
        exp_sheet.cell(row=2, column=7).value = 10
        exp_sheet.cell(row=3, column=7).value = 10
        exp_sheet.cell(row=4, column=7).value = 10
        exp_sheet.cell(row=5, column=7).value = 10
        exp_sheet.cell(row=6, column=7).value = 10
        exp_sheet.cell(row=7, column=7).value = 10
        exp_sheet.cell(row=1, column=1).font = Font(bold=True)
        exp_sheet.cell(row=1, column=2).font = Font(bold=True)
        exp_sheet.cell(row=1, column=3).font = Font(bold=True)
        exp_sheet.cell(row=1, column=4).font = Font(bold=True)
        exp_sheet.cell(row=1, column=5).font = Font(bold=True)
        exp_sheet.cell(row=1, column=6).font = Font(bold=True)
        exp_sheet.cell(row=1, column=7).font = Font(bold=True)

        for data in range(0, len(
                supp_parts_clean)):  # iterate through every cell in list and append the value in columns next to each other
            exp_sheet.cell(row=data + 2, column=1).value = supp_parts_clean[data]
            exp_sheet.cell(row=data + 2, column=2).value = COMPANY_parts[data][0]
            exp_sheet.cell(row=data + 2, column=3).value = qty_clean[data]
            exp_sheet.cell(row=data + 2, column=4).value = round(float(COMPANY_parts[data][1] * qty_clean[data]), 2)  # store purchasing price in column 4, rounded to 2 decimals
        os.chdir(os.path.dirname(filepath))
        label_error.destroy()
        if len(supp_parts) != 0:
            label_error2.destroy()
            save_file = asksaveasfilename(initialdir='.', title='Save File', filetypes=(('Excel files', (
        '*.xl*', '*.xlsx', '*.xlsm', '*.xlsb', '.xlam', '*.xltx', '*.xltm', '*.xls', '*.xla', '*.xlt', '*.xlm',
        '*.xlw')), ('all files', '*.*')))
            if save_file != "":
                export_excel.save(save_file + ".xlsx")
        else:
            label_error2.grid(row=2, column=0)

    except (NameError, FileNotFoundError):           # shows error message if no file is selected
        label_error.grid(row=2, column=0)

#Create window and GUI
window = tk.Tk()
window.title("part translator 10/2021")
window.columnconfigure(0, minsize = 300, weight = 1)
window.rowconfigure([0, 1, 2], minsize = 30, weight = 1)
file_select_button = tk.Button(text = "Select XML file from supplierx", command = open_file)
file_select_button.grid(row = 0, column = 0)
execute_button = tk.Button(text = "Save output as...", bg = "green", command = execute)
execute_button.grid(row = 1, column = 0)
label_error = tk.Label(text="Select input file", bg="red")
label_error2 = tk.Label(text="Incorrect XML input file selected", bg="red")

""" example part list database SUPPLIER_part : [COMPANY_part, purchase_price] """
parts_list = {
'p1':['c1', 38],
'p2':['c2', 355],
'p3':['c3', 35],
}

window.mainloop()
